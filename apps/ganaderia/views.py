# apps/ganaderia/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Animal, Pesaje, Parto, ProduccionLeche, EventoSalida
from .forms import PesajeForm, PartoForm, ProduccionForm, EventoSalidaForm, TrasladoForm, PesajeEditForm
from .services import AnimalService
from apps.users.decorators import role_required  # vamos a crear este decorador en users
from django.core.paginator import Paginator
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse, JsonResponse
from apps.ganaderia.models import Finca, Animal
from datetime import date


@login_required
def animales_list(request):
    q = request.GET.get('q')
    if q:
        animales = Animal.objects.filter(numero_arete__icontains=q) | Animal.objects.filter(nombre__icontains=q)
    else:
        animales = Animal.objects.all()
    return render(request, 'ganaderia/animales_list.html', {'animales': animales})

@login_required
def animal_detail(request, pk):
    animal = get_object_or_404(Animal, pk=pk)
    return render(request, 'ganaderia/animal_detail.html', {'animal': animal})

@login_required
def buscar_animal_por_arete(request):
    arete = request.GET.get("arete", "").strip()

    try:
        animal = Animal.objects.get_by_arete(arete)
        return JsonResponse({
            "existe": True,
            "nombre": animal.nombre,
            "finca": animal.finca.nombre,
        })
    except Animal.DoesNotExist:
        return JsonResponse({"existe": False})

@login_required
@role_required("Gerente", "Administrador Finca")
def pesajes_list_view(request):

    # --- REGISTRO DE PESAJE (POST) ---
    if request.method == "POST":
        form = PesajeForm(request.POST)

        if form.is_valid():
            try:
                AnimalService.registrar_pesaje(
                    numero_arete=form.cleaned_data['numero_arete'],
                    fecha=form.cleaned_data['fecha'],
                    peso=form.cleaned_data['peso'],
                    finca=form.cleaned_data['finca'],
                    usuario=request.user
                )

                messages.success(request, "Pesaje registrado exitosamente")
                return redirect('ganaderia:pesajes_list')

            except ValueError as e:
                messages.error(request, str(e))

    else:
        form = PesajeForm()

    # --- LISTADO DE PESAJES ---
    queryset = Pesaje.objects.all()

    # filtro por mes
    mes = request.GET.get('mes')
    if mes:
        queryset = queryset.filter(fecha__month=mes)

    # Exportar Excel
    if "exportar" in request.GET:
        return exportar_pesajes_excel(queryset)

    return render(request, 'ganaderia/pesajes_list.html', {
        'pesajes': queryset,
        'mes': mes,
        'form': form
    })

@login_required
@role_required("Gerente", "Administrador Finca")
def pesaje_edit_view(request, pk):
    pesaje = get_object_or_404(Pesaje, pk=pk)

    if request.method == 'POST':
        form = PesajeEditForm(request.POST, instance=pesaje)
        if form.is_valid():
            form.save()
            messages.success(request, "Pesaje actualizado correctamente.")
            return redirect('ganaderia:pesajes_list')
    else:
        form = PesajeEditForm(instance=pesaje)

    return render(request, 'ganaderia/pesaje_edit.html', {'form': form})


@login_required
@role_required("Gerente", "Administrador Finca")
def pesaje_delete_view(request, pk):
    pesaje = get_object_or_404(Pesaje, pk=pk)

    if request.method == "POST":
        pesaje.delete()
        messages.success(request, "Pesaje eliminado correctamente.")
        return redirect('ganaderia:pesajes_list')

    return render(request, 'ganaderia/pesaje_delete.html', {'pesaje': pesaje})


def exportar_pesajes_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pesajes"

    # Encabezados
    ws.append(["Fecha", "Arete", "Nombre", "Peso (kg)", "Finca"])

    # Datos
    for p in queryset:
        ws.append([
            p.fecha,
            p.animal.numero_arete,
            p.animal.nombre,
            p.peso,
            p.finca.nombre
        ])

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=pesajes.xlsx'

    wb.save(response)
    return response

@login_required
@role_required("Gerente", "Administrador Finca")
def partos_list(request):

    madre_filter = request.GET.get('madre', '')

    partos = Parto.objects.select_related("madre", "cria", "finca")

    if madre_filter:
        partos = partos.filter(madre__numero_arete__icontains=madre_filter)

    paginator = Paginator(partos.order_by('-fecha_nacimiento'), 10)
    page = request.GET.get('page')
    partos_page = paginator.get_page(page)

    context = {
        "partos": partos_page,
        "madre_filter": madre_filter
    }

    return render(request, "ganaderia/partos_list.html", context)

def editar_parto(request, id):
    parto = get_object_or_404(Parto, id=id)

    if request.method == 'POST':
        form = PartoForm(request.POST, instance=parto)
        if form.is_valid():
            form.save()
            return redirect('ganaderia:partos_list')
    else:
        form = PartoForm(instance=parto)

    return render(request, 'ganaderia/editar_parto.html', {'form': form})

def eliminar_parto(request, id):
    parto = get_object_or_404(Parto, id=id)
    parto.delete()
    return redirect('ganaderia:partos_list')


@login_required
@role_required("Gerente", "Administrador Finca")
def registrar_parto_view(request):

    if request.method == 'POST':

        madre_arete = request.POST.get("madre")
        fecha = request.POST.get("fecha_nacimiento")
        cria_arete = request.POST.get("numero_arete_cria")
        nombre_cria = request.POST.get("nombre_cria")
        sexo = request.POST.get("sexo")
        raza = request.POST.get("raza")
        peso = request.POST.get("peso")
        finca_id = request.POST.get("finca")

        # Validación madre
        try:
            madre = Animal.objects.get(numero_arete=madre_arete)
        except Animal.DoesNotExist:
            messages.error(request, "La madre no existe.")
            return redirect("ganaderia:registrar_parto")

        # Validación cría duplicada
        if Animal.objects.filter(numero_arete=cria_arete).exists():
            messages.error(request, "Ya existe un animal con ese número de arete.")
            return redirect("ganaderia:registrar_parto")

        # Crear cría
        cria = Animal.objects.create(
            numero_arete=cria_arete,
            nombre=nombre_cria,
            fecha_nacimiento=fecha,
            sexo=sexo,
            raza=raza,
            madre=madre,
            finca_id=finca_id,
            estado="activo"
        )

        # Registrar parto
        Parto.objects.create(
            fecha_nacimiento=fecha,
            madre=madre,
            cria=cria,
            finca_id=finca_id,
            peso=float(peso) if peso else None,
            raza=raza,
            sexo=sexo,
            created_by=request.user
        )

        messages.success(request, "Parto registrado exitosamente.")
        return redirect("ganaderia:partos_list")

    # GET → Mostrar formulario
    fincas = Finca.objects.all()
    madres = Animal.objects.filter(sexo="F")

    return render(request, "ganaderia/registrar_parto.html", {
        "fincas": fincas,
        "madres": madres
    })

@login_required
@role_required("Gerente", "Administrador Finca")
def partos_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Partos"

    # Encabezados
    ws.append([
        "Madre", "Fecha Nacimiento Cría", "Nombre Cría",
        "Arete Cría", "Finca", "Raza", "Sexo", "Peso (kg)"
    ])

    # Datos
    for p in Parto.objects.select_related("madre", "cria", "finca"):
        ws.append([
            p.madre.numero_arete,
            p.fecha_nacimiento.strftime("%d/%m/%Y"),
            p.cria.nombre,
            p.cria.numero_arete,
            p.finca.nombre,
            p.raza,
            p.sexo,
            p.peso,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="partos.xlsx"'
    wb.save(response)
    return response


@login_required
@role_required("Gerente", "Administrador Finca")
def registrar_produccion_view(request):
    mes = request.GET.get("mes")

    producciones = ProduccionLeche.objects.all()
    if mes:
        producciones = producciones.filter(fecha__month=mes)

    form_am = ProduccionForm(turno="AM")
    form_pm = ProduccionForm(turno="PM")

    meses = ["01","02","03","04","05","06","07","08","09","10","11","12"]

    return render(request, "ganaderia/registrar_produccion.html", {
        "form_am": form_am,
        "form_pm": form_pm,
        "producciones": producciones,
        "mes": mes,
        "meses": meses, 
    })

# ================================================================
#   REGISTRO PRODUCCIÓN AM
# ================================================================
def produccion_am_view(request):
    if request.method == "POST":
        form = ProduccionForm(request.POST)

        if form.is_valid():
            animal = form.cleaned_data["animal"]
            finca = form.cleaned_data["finca"]
            peso_am = form.cleaned_data["peso_am"]
            fecha = form.cleaned_data["fecha"]

            produccion, created = ProduccionLeche.objects.get_or_create(
                animal=animal,
                fecha=fecha,
                defaults={
                    "finca": finca,
                    "peso_am": peso_am,
                }
            )

            if not created:
                produccion.peso_am = peso_am
                produccion.save()

            messages.success(request, "Producción AM registrada correctamente.")
            return redirect("ganaderia:registrar_produccion")

    else:
        form = ProduccionForm()

    return render(request, "ganaderia/registrar_produccion.html", {"form": form})


# ================================================================
#   REGISTRO PRODUCCIÓN PM
# ================================================================
def produccion_pm_view(request):
    if request.method == "POST":
        form = ProduccionForm(request.POST)

        if form.is_valid():
            animal = form.cleaned_data["animal"]
            finca = form.cleaned_data["finca"]
            peso_pm = form.cleaned_data["peso_pm"]
            fecha = form.cleaned_data["fecha"]

            produccion, created = ProduccionLeche.objects.get_or_create(
                animal=animal,
                fecha=fecha,
                defaults={
                    "finca": finca,
                    "peso_pm": peso_pm,
                }
            )

            if not created:
                produccion.peso_pm = peso_pm
                produccion.save()

            messages.success(request, "Producción PM registrada correctamente.")
            return redirect("ganaderia:registrar_produccion")

    else:
        form = ProduccionForm()

    # 🔥 CORREGIDO: antes decía resgistrar_produccion.html
    return render(request, "ganaderia/registrar_produccion.html", {"form": form})



# ================================================================
#   EXPORTAR REGISTROS A EXCEL
# ================================================================
@login_required
@role_required("Gerente", "Administrador Finca")
def produccion_export_excel(request):

    mes = request.GET.get("mes")

    datos = ProduccionLeche.objects.all()
    if mes:
        datos = datos.filter(fecha__month=mes)

    wb = Workbook()
    ws = wb.active
    ws.title = "Producción"

    # Cabeceras
    ws.append(["Fecha", "Arete", "Nombre", "Finca", "AM", "PM", "Total Diario"])

    # Filas
    for p in datos:
        ws.append([
            p.fecha,
            p.animal.numero_arete,
            p.animal.nombre,
            p.finca.nombre,
            p.peso_am or 0,
            p.peso_pm or 0,
            p.total_diario,
        ])

    # Respuesta Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=produccion_leche.xlsx"
    wb.save(response)

    return response

@login_required
@role_required("Gerente", "Auxiliar administrativa")
def eventos_salida_list_view(request):

    eventos = EventoSalida.objects.select_related("animal").all()

    # --- FILTROS ---
    fecha = request.GET.get("fecha", "")
    tipo = request.GET.get("tipo", "")

    if fecha:
        eventos = eventos.filter(fecha=fecha)

    if tipo:
        eventos = eventos.filter(tipo_evento=tipo)

    return render(request, "ganaderia/eventos_salida_list.html", {
        "eventos": eventos,
        "fecha": fecha,
        "tipo": tipo,
    })

@login_required
@role_required("Gerente", "Auxiliar administrativa")
def registrar_evento_salida_view(request):

    if request.method == 'POST':
        form = EventoSalidaForm(request.POST)
        print("\n==== POST RECIBIDO ====")
        print(request.POST)

        if form.is_valid():
            print("==== FORM VALIDO ====")
            evento = form.save()
            print("==== EVENTO GUARDADO ====")
            print(evento.id, evento.animal_id, evento.tipo_evento)

            messages.success(request, "Evento de salida registrado exitosamente")
            return redirect('ganaderia:eventos_salida_list')
        else:
            print("==== ERRORES DE FORM ====")
            print(form.errors)

    else:
        form = EventoSalidaForm()

    return render(request, 'ganaderia/evento_salida_form.html', {'form': form})


@login_required
def api_animal_info(request):
    arete = request.GET.get("arete", "").strip()

    try:
        animal = Animal.objects.get(numero_arete=arete)
    except Animal.DoesNotExist:
        return JsonResponse({"error": "No encontrado"}, status=404)

    dias = (date.today() - animal.fecha_nacimiento).days

    return JsonResponse({
        "nombre": animal.nombre,
        "dias_nacido": dias
    })

@login_required
@role_required("Gerente", "Auxiliar administrativa")
def evento_salida_eliminar(request, id):
    evento = get_object_or_404(EventoSalida, id=id)
    evento.delete()
    messages.success(request, "Evento eliminado correctamente.")
    return redirect('ganaderia:eventos_salida_list')



@login_required
def traslados_view(request):

    # --- FILTRO POR ARETE ---
    q = request.GET.get("q", "")
    animales = Animal.objects.all().select_related("finca")

    if q:
        animales = animales.filter(numero_arete__icontains=q)

    fincas = Finca.objects.all()

    if request.method == "POST":
        aretes = request.POST.getlist("animales_seleccionados", [])
        finca_destino_id = request.POST.get("finca_destino")

        if not aretes:
            messages.error(request, "Debe seleccionar al menos un animal para realizar el traslado")
            return redirect("ganaderia:traslados")

        try:
            finca_destino = Finca.objects.get(id=finca_destino_id)
        except Finca.DoesNotExist:
            messages.error(request, "Debe seleccionar una finca destino válida")
            return redirect("ganaderia:traslados")

        try:
            AnimalService.trasladar_animales(aretes, finca_destino, usuario=request.user)
            messages.success(request, "Traslado realizado exitosamente")
            return redirect("ganaderia:traslados")
        except Exception as e:
            messages.error(request, str(e))

    return render(request, "ganaderia/traslados.html", {
        "animales": animales,
        "fincas": fincas,
        "q": q
    })
